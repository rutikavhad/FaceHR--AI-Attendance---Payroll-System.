from flask import Flask, render_template, request, redirect, url_for, flash
from flask_login import (
    LoginManager,
    login_user,
    logout_user,
    login_required,
    current_user
)
import numpy as np
import cv2
import pickle
import os
from sqlalchemy import extract

import pandas as pd
from io import BytesIO
from flask import send_file
import qrcode
from io import BytesIO
import base64
from insightface.app import FaceAnalysis
from flask import send_file
from openpyxl import Workbook
from io import BytesIO



from io import BytesIO
from flask import send_file
from openpyxl import Workbook
from sqlalchemy import or_
from model import *
from datetime import datetime, date
from sqlalchemy import func

app = Flask(__name__)

app.config['SECRET_KEY'] = 'secret-key'

app.config['SQLALCHEMY_DATABASE_URI'] = (
    'postgresql://matrix:1234@localhost:5432/company'
)

app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db.init_app(app)

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

face_app = FaceAnalysis(
    name='buffalo_l'
)

face_app.prepare(
    ctx_id=-1,
    det_size=(320, 320)
)


@login_manager.user_loader
def load_user(user_id):
    return Admin.query.get(int(user_id))


# @app.route('/')
# def home():

#     if current_user.is_authenticated:
#         return redirect(url_for('dashboard'))

#     return redirect(url_for('login'))


@app.route('/login', methods=['GET', 'POST'])
def login():

    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))

    if request.method == 'POST':

        email = request.form.get('email')
        password = request.form.get('password')

        admin = Admin.query.filter_by(email=email).first()

        if admin and admin.password == password:

            login_user(admin)

            flash('Login successful', 'success')

            return redirect(url_for('dashboard'))

        flash('Invalid email or password', 'danger')

    return render_template('login.html')

#main page

@app.route('/dashboard')
@login_required
def dashboard():

    today = date.today()

    total_employees = Employee.query.count()

    present = Attendance.query.filter_by(
        attendance_date=today,
        status='Present'
    ).count()

    half_day = Attendance.query.filter_by(
        attendance_date=today,
        status='Half Day'
    ).count()

    paid_leave = Attendance.query.filter_by(
        attendance_date=today,
        status='Paid Leave'
    ).count()

    unpaid_leave = Attendance.query.filter_by(
        attendance_date=today,
        status='Unpaid Leave'
    ).count()

    sick_leave = Attendance.query.filter_by(
        attendance_date=today,
        status='Sick Leave'
    ).count()

    absent = max(
        0,
        total_employees -
        present -
        half_day -
        paid_leave -
        unpaid_leave -
        sick_leave
    )

    return render_template(

        'dashboard.html',

        total_employees=total_employees,

        present=present,

        half_day=half_day,

        paid_leave=paid_leave,

        unpaid_leave=unpaid_leave,

        sick_leave=sick_leave,

        absent=absent
    )

from datetime import date, timedelta

@app.route('/attendance')
@login_required
def attendance_page():

    filter_type = request.args.get(
        'filter',
        'today'
    )

    start_date = request.args.get(
        'start_date'
    )

    end_date = request.args.get(
        'end_date'
    )

    search = request.args.get(
        'search',
        ''
    )

    query = db.session.query(
        Attendance,
        Employee
    ).join(
        Employee,
        Attendance.uid == Employee.uid
    )

    today = date.today()

    if filter_type == "today":

        query = query.filter(
            Attendance.attendance_date == today
        )

    elif filter_type == "week":

        week_start = today - timedelta(
            days=today.weekday()
        )

        query = query.filter(
            Attendance.attendance_date >= week_start
        )

    elif filter_type == "month":

        query = query.filter(
            func.extract(
                'month',
                Attendance.attendance_date
            ) == today.month,

            func.extract(
                'year',
                Attendance.attendance_date
            ) == today.year
        )

    elif filter_type == "year":

        query = query.filter(
            func.extract(
                'year',
                Attendance.attendance_date
            ) == today.year
        )

    elif (
        filter_type == "custom"
        and start_date
        and end_date
    ):

        query = query.filter(

            Attendance.attendance_date.between(
                start_date,
                end_date
            )

        )

    if search:

        query = query.filter(

            or_(

                Employee.name.ilike(
                    f'%{search}%'
                ),

                func.cast(
                    Employee.uid,
                    db.String
                ).ilike(
                    f'%{search}%'
                )

            )

        )

    records = query.order_by(
        Attendance.attendance_date.desc()
    ).all()

    return render_template(

        'attendence.html',

        records=records,

        filter_type=filter_type,

        start_date=start_date,

        end_date=end_date,

        search=search
    )    
    

    
@app.route('/manual-attendance')
@login_required
def manual_attendance():

    employees = Employee.query.all()

    today = date.today()

    attendance_map = {}

    records = Attendance.query.filter_by(
        attendance_date=today
    ).all()

    for row in records:

        attendance_map[row.uid] = row

    return render_template(

        'manual_attendance.html',

        employees=employees,

        attendance_map=attendance_map
    )
    
@app.route('/manual-checkin/<int:uid>')
@login_required
def manual_checkin(uid):

    employee = Employee.query.get(uid)

    if not employee:

        flash("Employee not found")

        return redirect(
            url_for('manual_attendance')
        )

    today = date.today()

    record = Attendance.query.filter_by(
        uid=uid,
        attendance_date=today
    ).first()

    if record:

        flash("Already checked in")

        return redirect(
            url_for('manual_attendance')
        )

    attendance = Attendance(

        uid=uid,

        attendance_date=today,

        login_time=datetime.now(),

        status="Present"
    )

    try:

        db.session.add(attendance)

        db.session.commit()

        flash("Check-In Successful")

    except Exception as e:

        db.session.rollback()

        flash(str(e))

    return redirect(
        url_for('manual_attendance')
    )

@app.route('/manual-checkout/<int:uid>')
@login_required
def manual_checkout(uid):

    today = date.today()

    attendance = Attendance.query.filter_by(
        uid=uid,
        attendance_date=today
    ).first()

    if not attendance:

        flash("No attendance record found")

        return redirect(
            url_for('manual_attendance')
        )

    if attendance.login_time is None:

        flash("Employee has not checked in yet")

        return redirect(
            url_for('manual_attendance')
        )

    if attendance.logout_time:

        flash("Already checked out")

        return redirect(
            url_for('manual_attendance')
        )

    logout_time = datetime.now()

    attendance.logout_time = logout_time

    total = logout_time - attendance.login_time

    attendance.total_work_time = total

    hours = total.total_seconds() / 3600

    if hours < 4:
        attendance.status = "Half Day"
    else:
        attendance.status = "Present"

    db.session.commit()

    flash("Check-Out Successful")

    return redirect(
        url_for('manual_attendance')
    )
    
    
    
    
    
@app.route('/manual-attendance/save', methods=['POST'])
@login_required
def manual_attendance_save():

    uid = request.form['uid']

    status = request.form['status']

    remarks = request.form['remarks']

    today = date.today()

    record = Attendance.query.filter_by(
        uid=uid,
        attendance_date=today
    ).first()

    if record:

        record.status = status

        record.remarks = remarks

    else:

        record = Attendance(

            uid=uid,

            attendance_date=today,

            login_time=datetime.now(),

            logout_time=datetime.now(),

            total_work_time=None,

            status=status,

            remarks=remarks
        )

        db.session.add(record)

    db.session.commit()

    flash("Attendance saved.")

    return redirect(
        url_for('manual_attendance')
    )     


    

@app.route('/attendance/checkin/<int:uid>')
@login_required
def checkin(uid):

    employee = Employee.query.get(uid)

    if not employee:

        flash("Employee not found")

        return redirect(url_for('attendance_page'))

    today = date.today()

    attendance = Attendance.query.filter_by(
        uid=uid,
        attendance_date=today
    ).first()

    if attendance:

        flash("Already checked in.")

        return redirect(url_for('attendance_page'))

    attendance = Attendance(

        uid=uid,

        attendance_date=today,

        login_time=datetime.now(),

        status="Present"
    )

    db.session.add(attendance)

    db.session.commit()

    flash(
        f"{employee.name} checked in successfully."
    )

    return redirect(url_for('attendance_page'))


@app.route('/attendance/checkout/<int:uid>')
@login_required
def checkout(uid):

    today = date.today()

    attendance = Attendance.query.filter_by(
        uid=uid,
        attendance_date=today
    ).first()

    if not attendance:

        flash("Check-in first.")

        return redirect(url_for('attendance_page'))

    if attendance.logout_time:

        flash("Already checked out.")

        return redirect(url_for('attendance_page'))

    logout_time = datetime.now()

    attendance.logout_time = logout_time

    total = logout_time - attendance.login_time

    attendance.total_work_time = total

    hours = total.total_seconds() / 3600

    if hours < 4:
        attendance.status = "Half Day"
    else:
        attendance.status = "Present"

    db.session.commit()

    flash("Check-out successful.")

    return redirect(url_for('attendance_page'))# from sqlalchemy import or_

@app.route('/employees')
@login_required
def employees():

    search = request.args.get('search', '')

    query = Employee.query

    if search:

        query = query.filter(

            or_(

                Employee.name.ilike(f'%{search}%'),

                Employee.position.ilike(f'%{search}%'),

                Employee.phone.ilike(f'%{search}%')

            )

        )

    employees = query.all()

    return render_template(
        'employee.html',
        employees=employees
    )    

@app.route('/employee/add', methods=['GET', 'POST'])
@login_required
def add_employee():

    if request.method == 'POST':

        employee = Employee(

            name=request.form['name'],

            phone=request.form['phone'],

            addr=request.form['addr'],

            addhar=request.form['addhar'],

            position=request.form['position'],

            monthly_salary=request.form[
                'monthly_salary'
            ],

            bank_account=request.form[
                'bank_account'
            ],

            branch_name=request.form[
                'branch_name'
            ],

            ifsc_code=request.form[
                'ifsc_code'
            ],

            upi_id=request.form[
                'upi_id'
            ]
        )

        db.session.add(employee)

        db.session.commit()

        return redirect(url_for('employees'))

    return render_template('employee_form.html')

@app.route('/employee/edit/<int:uid>', methods=['GET', 'POST'])
@login_required
def edit_employee(uid):

    employee = Employee.query.get_or_404(uid)

    if request.method == 'POST':

        employee.name = request.form['name']

        employee.phone = request.form['phone']

        employee.addr = request.form['addr']

        employee.addhar = request.form['addhar']

        employee.position = request.form['position']

        employee.monthly_salary = request.form[
            'monthly_salary'
        ]

        employee.bank_account = request.form[
            'bank_account'
        ]

        employee.branch_name = request.form[
            'branch_name'
        ]

        employee.ifsc_code = request.form[
            'ifsc_code'
        ]

        employee.upi_id = request.form[
            'upi_id'
        ]

        db.session.commit()

        return redirect(url_for('employees'))

    return render_template(
        'employee_form.html',
        employee=employee
    )


@app.route('/employee/delete/<int:uid>')
@login_required
def delete_employee(uid):

    employee = Employee.query.get_or_404(uid)

    db.session.delete(employee)

    db.session.commit()

    flash('Employee deleted')

    return redirect(url_for('employees'))


@app.route('/logout')
@login_required
def logout():

    logout_user()

    flash('Logged out successfully')

    return redirect(url_for('login'))



@app.route('/reports')
@login_required
def reports():

    return render_template('report.html')


@app.route('/reports/employees/export')
@login_required
def export_employees():

    employees = Employee.query.all()

    wb = Workbook()

    ws = wb.active

    ws.title = "Employees"

    ws.append([

        "UID",

        "Name",

        "Phone",

        "Address",

        "Aadhaar",

        "Position"

    ])

    for employee in employees:

        ws.append([

            employee.uid,

            employee.name,

            employee.phone,

            employee.addr,

            employee.addhar,

            employee.position

        ])

    output = BytesIO()

    wb.save(output)

    output.seek(0)

    return send_file(

        output,

        download_name='employees.xlsx',

        as_attachment=True,

        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    


@app.route('/reports/attendance/export')
@login_required
def export_attendance():

    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    if not start_date or not end_date:

        flash("Please select dates")

        return redirect(url_for('report'))

    records = db.session.query(

        Attendance,
        Employee

    ).join(

        Employee,
        Attendance.uid == Employee.uid

    ).filter(

        Attendance.attendance_date.between(
            start_date,
            end_date
        )

    ).all()

    wb = Workbook()

    ws = wb.active

    ws.title = "Attendance"

    headers = [

        "Attendance ID",
        "UID",
        "Employee Name",
        "Phone",
        "Aadhaar",
        "Position",
        "Attendance Date",
        "Login Time",
        "Logout Time",
        "Total Work Time",
        "Status",
        "Remarks"

    ]

    ws.append(headers)

    for attendance, employee in records:

        ws.append([

            attendance.attendance_id,

            employee.uid,

            employee.name,

            employee.phone,

            employee.addhar,

            employee.position,

            str(attendance.attendance_date),

            str(attendance.login_time),

            str(attendance.logout_time),

            attendance.total_work_time,

            attendance.status,

            attendance.remarks

        ])

    output = BytesIO()

    wb.save(output)

    output.seek(0)

    return send_file(

        output,

        download_name='attendance_report.xlsx',

        as_attachment=True,

        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )



@app.route('/leave')
@login_required
def leave():

    today = date.today()

    records = db.session.query(
        Employee
    ).all()

    return render_template(
        'leave.html',
        records=records,
        today=today
    )
    
@app.route('/leave/mark/<int:uid>', methods=['POST'])
@login_required
def mark_leave(uid):

    attendance_date = request.form['attendance_date']

    status = request.form['status']

    remarks = request.form['remarks']

    record = Attendance.query.filter_by(
        uid=uid,
        attendance_date=attendance_date
    ).first()

    if record:

        record.status = status
        record.remarks = remarks

    else:

        record = Attendance(

            uid=uid,

            attendance_date=attendance_date,

            status=status,

            remarks=remarks
        )

        db.session.add(record)

    db.session.commit()

    flash("Leave updated successfully.")

    return redirect(url_for('leave'))



@app.route('/employee-summary')
@login_required
def employee_summary():

    summaries = []

    employees = Employee.query.all()

    for employee in employees:

        present = Attendance.query.filter_by(
            uid=employee.uid,
            status='Present'
        ).count()

        half_day = Attendance.query.filter_by(
            uid=employee.uid,
            status='Half Day'
        ).count()

        paid_leave = Attendance.query.filter_by(
            uid=employee.uid,
            status='Paid Leave'
        ).count()

        unpaid_leave = Attendance.query.filter_by(
            uid=employee.uid,
            status='Unpaid Leave'
        ).count()

        sick_leave = Attendance.query.filter_by(
            uid=employee.uid,
            status='Sick Leave'
        ).count()

        summaries.append({

            'employee': employee,

            'present': present,

            'half_day': half_day,

            'paid_leave': paid_leave,

            'unpaid_leave': unpaid_leave,

            'sick_leave': sick_leave
        })

    return render_template(

        'employee_summary.html',

        summaries=summaries
    )


@app.route('/reports/summary/export')
@login_required
def export_summary():

    wb = Workbook()

    ws = wb.active

    ws.title = 'Summary'

    ws.append([

        'UID',

        'Name',

        'Position',

        'Present',

        'Half Day',

        'Paid Leave',

        'Unpaid Leave',

        'Sick Leave'

    ])

    employees = Employee.query.all()

    for employee in employees:

        ws.append([

            employee.uid,

            employee.name,

            employee.position,

            Attendance.query.filter_by(
                uid=employee.uid,
                status='Present'
            ).count(),

            Attendance.query.filter_by(
                uid=employee.uid,
                status='Half Day'
            ).count(),

            Attendance.query.filter_by(
                uid=employee.uid,
                status='Paid Leave'
            ).count(),

            Attendance.query.filter_by(
                uid=employee.uid,
                status='Unpaid Leave'
            ).count(),

            Attendance.query.filter_by(
                uid=employee.uid,
                status='Sick Leave'
            ).count()

        ])

    output = BytesIO()

    wb.save(output)

    output.seek(0)

    return send_file(

        output,

        download_name='employee_summary.xlsx',

        as_attachment=True,

        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
#for auto attendence by face
#####################################################################################################################################
def do_checkin(uid):

    today = date.today()

    attendance = Attendance.query.filter_by(
        uid=uid,
        attendance_date=today
    ).first()

    if attendance:

        return False, "Already checked in"

    attendance = Attendance(
        uid=uid,
        attendance_date=today,
        login_time=datetime.now(),
        status="Present"
    )

    db.session.add(attendance)

    db.session.commit()

    return True, "Check-In Successful"

def do_checkout(uid):

    today = date.today()

    attendance = Attendance.query.filter_by(
        uid=uid,
        attendance_date=today
    ).first()

    if not attendance:

        return False, "Check-in first"

    if attendance.logout_time:

        return False, "Already checked out"

    logout_time = datetime.now()

    attendance.logout_time = logout_time

    total = logout_time - attendance.login_time

    attendance.total_work_time = total

    hours = total.total_seconds() / 3600

    if hours < 4:
        attendance.status = "Half Day"
    else:
        attendance.status = "Present"

    db.session.commit()

    return True, "Check-Out Successful"


#####################################################################################################################################
@app.route('/register-face')
@login_required
def register_face_list():

    employees = Employee.query.all()

    return render_template(
        'register_face_list.html',
        employees=employees
    )

@app.route('/register-face/<int:uid>')
@login_required
def register_face(uid):

    employee = Employee.query.get_or_404(uid)

    return render_template(
        'register_face.html',
        employee=employee
    )


@app.route('/api/register-face/<int:uid>')
@login_required
def api_register_face(uid):

    employee = Employee.query.get_or_404(uid)

    db_file = "faces.pkl"

    database = {}

    if os.path.exists(db_file):

        try:

            with open(db_file, "rb") as f:

                database = pickle.load(f)

        except Exception:

            database = {}
    cap = cv2.VideoCapture(0)

    while True:

        ret, frame = cap.read()

        if not ret:
            break

        faces = face_app.get(frame)

        for face in faces:

            x1, y1, x2, y2 = (
                face.bbox.astype(int)
            )

            cv2.rectangle(

                frame,

                (x1, y1),

                (x2, y2),

                (0,255,0),

                2
            )

            cv2.putText(

                frame,

                employee.name,

                (x1, y1-10),

                cv2.FONT_HERSHEY_SIMPLEX,

                0.8,

                (0,255,0),

                2
            )

        cv2.imshow(
            "Register Face",
            frame
        )

        key = cv2.waitKey(1)

        if key == 13:

            if len(faces) != 1:

                continue

            database[uid] = {

                "uid": uid,

                "name": employee.name,

                "embedding":
                    faces[0].embedding
            }

            try:

                with open(db_file, "wb") as f:

                    pickle.dump(database, f)

            except Exception as e:

                cap.release()

                cv2.destroyAllWindows()

                return {
                    "success": False,
                    "message": str(e)
                }

            cap.release()

            cv2.destroyAllWindows()

            return {

                "success": True,

                "message":
                    f"{employee.name} registered."
            }

        elif key == 27:

            cap.release()

            cv2.destroyAllWindows()

            return {

                "success": False,

                "message":
                    "Cancelled."
            }

def cosine_similarity(a,b):

    return np.dot(a,b)/(

        np.linalg.norm(a)

        *

        np.linalg.norm(b)

    )
    
@app.route('/kiosk')
def kiosk():

    return render_template(
        'kiosk.html'
    )

@app.route('/api/face-attendance')
def face_attendance():

    try:

        db_file = "faces.pkl"

        if not os.path.exists(db_file):

            return {
                "success": False,
                "message": "No registered faces found."
            }

        with open(db_file, "rb") as f:

            database = pickle.load(f)

        if len(database) == 0:

            return {
                "success": False,
                "message": "Face database is empty."
            }

        cap = cv2.VideoCapture(0)

        THRESHOLD = 0.75

        while True:

            ret, frame = cap.read()

            if not ret:
                continue

            faces = face_app.get(frame)

            for face in faces:

                emb = face.embedding

                best_uid = None
                best_score = -1

                for uid, data in database.items():

                    score = cosine_similarity(

                        emb,

                        data["embedding"]

                    )

                    print(
                        f"UID={uid} SCORE={score:.4f}"
                    )

                    if score > best_score:

                        best_score = score
                        best_uid = uid

                x1, y1, x2, y2 = (
                    face.bbox.astype(int)
                )

                if best_score < THRESHOLD:

                    cv2.rectangle(

                        frame,

                        (x1, y1),

                        (x2, y2),

                        (0, 0, 255),

                        2

                    )

                    cv2.putText(

                        frame,

                        f"Unknown ({best_score:.2f})",

                        (x1, y1 - 10),

                        cv2.FONT_HERSHEY_SIMPLEX,

                        0.7,

                        (0, 0, 255),

                        2

                    )

                    continue

                employee = Employee.query.get(
                    int(best_uid)
                )

                if not employee:
                    continue

                cv2.rectangle(

                    frame,

                    (x1, y1),

                    (x2, y2),

                    (0, 255, 0),

                    2

                )

                cv2.putText(

                    frame,

                    f"{employee.name} ({best_score:.2f})",

                    (x1, y1 - 10),

                    cv2.FONT_HERSHEY_SIMPLEX,

                    0.7,

                    (0, 255, 0),

                    2

                )

                cv2.imshow(
                    "Attendance",
                    frame
                )

                key = cv2.waitKey(1)

                if key == 13:

                    today = date.today()

                    attendance = Attendance.query.filter_by(

                        uid=employee.uid,

                        attendance_date=today

                    ).first()

                    now = datetime.now()

                    # CHECK-IN

                    if attendance is None:

                        attendance = Attendance(

                            uid=employee.uid,

                            attendance_date=today,

                            login_time=now,

                            status="Present"

                        )

                        db.session.add(
                            attendance
                        )

                        db.session.commit()

                        cap.release()

                        cv2.destroyAllWindows()

                        return {

                            "success": True,

                            "employee": employee.name,

                            "action": "checkin",

                            "message":
                                f"{employee.name} Check-In Successful"

                        }

                    # CHECK-OUT

                    elif attendance.logout_time is None:

                        attendance.logout_time = now

                        total = (

                            attendance.logout_time

                            - attendance.login_time

                        )

                        attendance.total_work_time = total

                        db.session.commit()

                        cap.release()

                        cv2.destroyAllWindows()

                        return {

                            "success": True,

                            "employee": employee.name,

                            "action": "checkout",

                            "message":
                                f"{employee.name} Check-Out Successful"

                        }

                    # ALREADY DONE

                    else:

                        cap.release()

                        cv2.destroyAllWindows()

                        return {

                            "success": False,

                            "employee": employee.name,

                            "message":
                                "Attendance already completed."

                        }

            cv2.imshow(
                "Attendance",
                frame
            )

            if cv2.waitKey(1) == 27:

                break

        cap.release()

        cv2.destroyAllWindows()

        return {

            "success": False,

            "message": "No face detected."

        }

    except Exception as e:

        print("FACE ERROR:", e)

        return {

            "success": False,

            "message": str(e)

        }, 500


@app.route('/api/do-logout/<int:uid>', methods=['POST'])
def do_logout_api(uid):

    success, message = do_checkout(uid)

    return {
        "success": success,
        "message": message
    }
    
    
THRESHOLD = 0.6

def cosine_similarity(a, b):

    return np.dot(a, b) / (

        np.linalg.norm(a) *

        np.linalg.norm(b)

    )


@app.route('/api/recognize', methods=['POST'])
def recognize():

    try:

        # TEMP TEST
        uid = 1

        employee = Employee.query.get(uid)

        if not employee:

            return {
                "message": "Employee not found"
            }

        today = date.today()

        attendance = Attendance.query.filter_by(
            uid=uid,
            attendance_date=today
        ).first()

        if attendance is None:

            action = "login"

            login_time = ""

            logout_time = ""

        elif attendance.logout_time is None:

            action = "logout"

            login_time = str(
                attendance.login_time
            )

            logout_time = ""

        else:

            action = "completed"

            login_time = str(
                attendance.login_time
            )

            logout_time = str(
                attendance.logout_time
            )

        return {

            "uid": employee.uid,

            "name": employee.name,

            "action": action,

            "login_time": login_time,

            "logout_time": logout_time

        }

    except Exception as e:

        print(e)

        return {
            "message": str(e)
        }, 500
        


def do_checkin(uid):

    today = date.today()

    attendance = Attendance.query.filter_by(
        uid=uid,
        attendance_date=today
    ).first()

    if attendance:

        return False, "Already checked in"

    attendance = Attendance(
        uid=uid,
        attendance_date=today,
        login_time=datetime.now(),
        status="Present"
    )

    db.session.add(attendance)
    db.session.commit()

    return True, "Check-In Successful"

def do_checkout(uid):

    today = date.today()

    attendance = Attendance.query.filter_by(
        uid=uid,
        attendance_date=today
    ).first()

    if not attendance:

        return False, "Check-in first"

    if attendance.logout_time:

        return False, "Already checked out"

    logout_time = datetime.now()

    attendance.logout_time = logout_time

    total = logout_time - attendance.login_time

    attendance.total_work_time = total

    hours = total.total_seconds() / 3600

    if hours < 4:
        attendance.status = "Half Day"
    else:
        attendance.status = "Present"

    db.session.commit()

    return True, "Check-Out Successful"


@app.route('/api/do-login/<int:uid>', methods=['POST'])
def do_login_api(uid):

    success, message = do_checkin(uid)

    return {
        "success": success,
        "message": message
    }

##################################################################################################################
#pay option
@app.route('/payroll')
@login_required
def payroll():

    today = date.today()

    month = request.args.get(
        'month',
        today.month,
        type=int
    )

    year = request.args.get(
        'year',
        today.year,
        type=int
    )

    search = request.args.get(
        'search',
        ''
    )

    results = []

    employees_query = Employee.query

    if search:

        employees_query = employees_query.filter(

            or_(

                Employee.name.ilike(
                    f"%{search}%"
                ),

                func.cast(
                    Employee.uid,
                    db.String
                ).ilike(
                    f"%{search}%"
                )

            )

        )

    employees = employees_query.all()

    for emp in employees:

        records = Attendance.query.filter(

            Attendance.uid == emp.uid,

            extract(
                'month',
                Attendance.attendance_date
            ) == month,

            extract(
                'year',
                Attendance.attendance_date
            ) == year

        ).all()

        total_hours = 0
        total_days = 0
        paid_leave_days = 0

        for r in records:

            if r.status == "Present":

                if r.total_work_time:

                    hours = (
                        r.total_work_time.total_seconds()
                        / 3600
                    )

                    hours = min(hours, 9)

                else:

                    hours = 9

                total_hours += hours
                total_days += 1

            elif r.status == "Paid Leave":

                total_hours += 9
                total_days += 1
                paid_leave_days += 1

        monthly_salary = float(
            emp.monthly_salary or 0
        )

        per_day_salary = monthly_salary / 26

        per_hour_salary = per_day_salary / 9

        calculated_salary = round(

            total_hours *
            per_hour_salary,

            2

        )

        payment = Payment.query.filter_by(

            uid=emp.uid,

            pay_month=month,

            pay_year=year

        ).first()

        if payment:

            payment_status = payment.payment_status

            payment_date = payment.payment_date

            payment_remarks = payment.remarks

        else:

            payment_status = "Pending"

            payment_date = None

            payment_remarks = ""

        results.append({

            "employee": emp,

            "days": total_days,

            "hours": round(
                total_hours,
                2
            ),

            "paid_leave_days":
                paid_leave_days,

            "monthly_salary":
                monthly_salary,

            "calculated_salary":
                calculated_salary,

            "payment_status":
                payment_status,

            "payment_date":
                payment_date,

            "payment_remarks":
                payment_remarks

        })

    return render_template(

        "payroll.html",

        results=results,

        month=month,

        year=year,

        search=search
    )
        
@app.route('/payment/<int:uid>')
@login_required
def payment_page(uid):

    employee = Employee.query.get_or_404(uid)

    month = request.args.get(
        "month",
        date.today().month,
        type=int
    )

    year = request.args.get(
        "year",
        date.today().year,
        type=int
    )

    records = Attendance.query.filter(

        Attendance.uid == uid,

        extract(
            'month',
            Attendance.attendance_date
        ) == month,

        extract(
            'year',
            Attendance.attendance_date
        ) == year

    ).all()

    total_hours = 0
    total_days = 0
    paid_leave_days = 0

    for r in records:

        if r.status == "Present":

            if r.total_work_time:

                hours = (
                    r.total_work_time.total_seconds()
                    / 3600
                )

                hours = min(hours, 9)

            else:

                hours = 9

            total_hours += hours
            total_days += 1

        elif r.status == "Paid Leave":

            total_hours += 9
            total_days += 1
            paid_leave_days += 1

    monthly_salary = float(
        employee.monthly_salary or 0
    )

    per_day_salary = (
        monthly_salary / 26
    )

    per_hour_salary = (
        per_day_salary / 9
    )

    salary = round(

        total_hours *
        per_hour_salary,

        2

    )

    upi_link = (

        f"upi://pay"

        f"?pa={employee.upi_id}"

        f"&pn={employee.name}"

        f"&am={salary}"

        f"&cu=INR"

    )

    qr = qrcode.make(
        upi_link
    )

    buffer = BytesIO()

    qr.save(
        buffer,
        format="PNG"
    )

    qr_base64 = base64.b64encode(
        buffer.getvalue()
    ).decode()

    return render_template(

        "payment_page.html",

        employee=employee,

        month=month,

        year=year,

        salary=salary,

        total_days=total_days,

        total_hours=round(
            total_hours,
            2
        ),

        paid_leave_days=
            paid_leave_days,

        qr_code=qr_base64
    )

@app.route(
    '/payment/confirm/<int:uid>',
    methods=['POST']
)
@login_required
def confirm_payment(uid):
    existing = Payment.query.filter_by(

        uid=uid,

        pay_month=date.today().month,

        pay_year=date.today().year

    ).first()

    if existing:

        flash(
            "Salary already paid"
        )

        return redirect(
            url_for('payroll')
        )

    employee = Employee.query.get_or_404(uid)

    salary = float(
        request.form['salary']
    )

    payment = Payment(

        uid=uid,

        pay_month=date.today().month,

        pay_year=date.today().year,

        paid_amount=salary,

        payment_date=datetime.now(),

        payment_status="Paid",

        bank_account=employee.bank_account,

        branch_name=employee.branch_name,

        remarks="Salary Paid"
    )

    db.session.add(payment)

    db.session.commit()

    flash(
        "Payment marked successfully"
    )

    return redirect(
        url_for('payroll')
    )
    

@app.route('/payroll/excel')
@login_required
def payroll_excel():

    month = request.args.get(
        'month',
        date.today().month,
        type=int
    )

    year = request.args.get(
        'year',
        date.today().year,
        type=int
    )

    search = request.args.get(
        'search',
        ''
    )

    employees_query = Employee.query

    if search:

        employees_query = employees_query.filter(

            or_(

                Employee.name.ilike(
                    f"%{search}%"
                ),

                func.cast(
                    Employee.uid,
                    db.String
                ).ilike(
                    f"%{search}%"
                )

            )

        )

    employees = employees_query.all()

    rows = []

    for emp in employees:

        records = Attendance.query.filter(

            Attendance.uid == emp.uid,

            extract(
                'month',
                Attendance.attendance_date
            ) == month,

            extract(
                'year',
                Attendance.attendance_date
            ) == year

        ).all()

        total_hours = 0
        total_days = 0
        paid_leave_days = 0
        present_days = 0

        for r in records:

            if r.status == "Present":

                if r.total_work_time:

                    hours = (
                        r.total_work_time.total_seconds()
                        / 3600
                    )

                    hours = min(
                        hours,
                        9
                    )

                else:

                    hours = 9

                total_hours += hours
                total_days += 1
                present_days += 1

            elif r.status == "Paid Leave":

                total_hours += 9
                total_days += 1
                paid_leave_days += 1

        monthly_salary = float(
            emp.monthly_salary or 0
        )

        per_day_salary = (
            monthly_salary / 26
        )

        per_hour_salary = (
            per_day_salary / 9
        )

        calculated_salary = round(

            total_hours *
            per_hour_salary,

            2

        )

        payment = Payment.query.filter_by(

            uid=emp.uid,

            pay_month=month,

            pay_year=year

        ).first()

        rows.append({

            "UID":
                emp.uid,

            "Name":
                emp.name,

            "Phone":
                emp.phone,

            "Position":
                emp.position,

            "Bank Account":
                emp.bank_account,

            "Branch":
                emp.branch_name,

            "IFSC":
                emp.ifsc_code,

            "UPI":
                emp.upi_id,

            "Month":
                month,

            "Year":
                year,

            "Present Days":
                present_days,

            "Paid Leave":
                paid_leave_days,

            "Total Days":
                total_days,

            "Total Hours":
                round(
                    total_hours,
                    2
                ),

            "Monthly Salary":
                monthly_salary,

            "Per Day Salary":
                round(
                    per_day_salary,
                    2
                ),

            "Per Hour Salary":
                round(
                    per_hour_salary,
                    2
                ),

            "Calculated Salary":
                calculated_salary,

            "Payment Status":
                payment.payment_status
                if payment
                else "Pending",

            "Payment Date":
                payment.payment_date
                if payment
                else "",

            "Remarks":
                payment.remarks
                if payment
                else ""

        })

    df = pd.DataFrame(rows)

    output = BytesIO()

    with pd.ExcelWriter(

        output,

        engine='openpyxl'

    ) as writer:

        df.to_excel(

            writer,

            sheet_name='Payroll',

            index=False

        )

    output.seek(0)

    return send_file(

        output,

        as_attachment=True,

        download_name=
        f'Payroll_{month}_{year}.xlsx',

        mimetype=
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


if __name__ == '__main__':
    app.run(debug=True)